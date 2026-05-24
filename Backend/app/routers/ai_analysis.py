# Backend/app/routers/ai_analysis.py
from fastapi import APIRouter, HTTPException
import httpx
import os
from dotenv import load_dotenv
import zipfile
import openpyxl
import pandas as pd
from pathlib import Path
import logging
import sys

# Load environment variables from .env if present
load_dotenv()

# Setup logging with Windows-friendly encoding
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

router = APIRouter()

# Ollama settings: allow overriding via .env (e.g. OLLAMA_URL=http://ollama:11434)
_ollama_base = os.getenv("OLLAMA_URL")
if not _ollama_base:
    logger.warning("OLLAMA_URL not set, using default")
    _ollama_base = "http://ollama:11434"
    
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.2")
OLLAMA_URL = f"{_ollama_base.rstrip('/')}/api/generate"

# Import from config - this will now use the Docker-friendly paths
try:
    from app.config import REJECTED_DIR_STR as REJECTED_DIR
    logger.info(f"REJECTED_DIR loaded from config: {REJECTED_DIR}")
except ImportError as e:
    logger.error(f"Failed to import REJECTED_DIR from config: {e}")
    # Fallback for Windows
    REJECTED_DIR = os.getenv("REJECTED_DIR", "/app/etl/rejected")
    logger.info(f"Using fallback REJECTED_DIR: {REJECTED_DIR}")

# Log the rejected directory path for debugging
logger.info(f"REJECTED_DIR path: {REJECTED_DIR}")
logger.info(f"Directory exists: {os.path.exists(REJECTED_DIR)}")
logger.info(f"Current working directory: {os.getcwd()}")
logger.info(f"Directory contents of /app: {os.listdir('/app') if os.path.exists('/app') else 'N/A'}")
logger.info(f"Directory contents of /app/etl: {os.listdir('/app/etl') if os.path.exists('/app/etl') else 'N/A'}")

# Business rules that define mandatory sheets and columns for validation
BUSINESS_RULES = """
Each uploaded Excel file must contain two mandatory sheets: 'OP' and 'DEVIS'.

Sheet OP - Required columns:
- Oppo_OpportunityId   : unique opportunity ID
- Oppo_CreatedDate     : date the opportunity was created
- Oppo_AssignedUserId  : ID of the assigned salesperson
- User_LastName        : salesperson last name
- User_FirstName       : salesperson first name
- User_EmailAddress    : salesperson email
- Chan_Description     : agency/channel name
- Comp_Name            : client company name

Sheet OP - Optional (missing = warning only, not rejection):
- Oppo_Deleted, Addr_City, Emai_EmailAddress

Sheet DEVIS - Required columns:
- Quot_opportunityid   : links quote to an opportunity
- Quot_OrderQuoteID    : unique quote ID
- AR_Ref               : vehicle reference (primary key)
- User_LastName        : salesperson last name
- User_FirstName       : salesperson first name

Sheet DEVIS - Optional (missing = warning only):
- User_UserId, Quot_CreatedDate, AR_Design, Marque, Modele

Sheet SALES - Fully optional. If present, expected columns:
quote_id, oppo_id, ar_ref, user_id, sale_date, quantity, final_price

A file is REJECTED when:
- Sheet 'OP' or 'DEVIS' is completely missing
- Either mandatory sheet is empty
- Any required column listed above is absent from its sheet
"""


# Attempt to read Excel headers using openpyxl
def _read_with_openpyxl(fpath):
    """Read Excel file using openpyxl (standard method)"""
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c) if c is not None else "" for c in row]
                break
            sheets[sheet_name] = headers
        wb.close()
        logger.info(f"Successfully read {fpath} with openpyxl")
        return sheets
    except Exception as e:
        logger.warning(f"openpyxl failed for {fpath}: {str(e)}")
        raise


# Fallback reader using pandas to extract sheet headers
def _read_with_pandas(fpath):
    """Read Excel file using pandas as fallback"""
    try:
        sheets = {}
        # Try different engines for Windows compatibility
        engines = ["openpyxl", "xlrd"]
        for engine in engines:
            try:
                xl = pd.ExcelFile(fpath, engine=engine)
                for sheet_name in xl.sheet_names:
                    df = xl.parse(sheet_name, nrows=0)
                    sheets[sheet_name] = [str(col) for col in df.columns]
                logger.info(f"Successfully read {fpath} with pandas (engine={engine})")
                return sheets
            except Exception as e:
                logger.warning(f"pandas with engine {engine} failed: {str(e)}")
                continue
        raise Exception("All pandas engines failed")
    except Exception as e:
        logger.warning(f"pandas failed for {fpath}: {str(e)}")
        raise


# Last-resort reader that opens the xlsx as a zip to list sheets
def _read_as_zip(fpath):
    """Last resort: open as raw zip to at least get sheet names"""
    try:
        sheets = {}
        with zipfile.ZipFile(fpath, "r") as z:
            for name in z.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    sheet_id = name.split("/")[-1].replace(".xml", "")
                    sheets[f"sheet_{sheet_id}"] = []
        logger.info(f"Partially read {fpath} as zip archive")
        return sheets
    except Exception as e:
        logger.error(f"Zip read failed for {fpath}: {str(e)}")
        raise


# Read and summarize all rejected Excel files for AI analysis
def read_rejected_files():
    """Read all rejected Excel files from the rejected directory"""
    results = []
    
    # Create directory if it doesn't exist (for Docker on Windows)
    if not os.path.exists(REJECTED_DIR):
        logger.warning(f"REJECTED_DIR does not exist: {REJECTED_DIR}")
        try:
            os.makedirs(REJECTED_DIR, exist_ok=True)
            logger.info(f"Created REJECTED_DIR: {REJECTED_DIR}")
            return results  # Return empty since directory is new
        except Exception as e:
            logger.error(f"Failed to create REJECTED_DIR: {str(e)}")
            return results
    
    # List all files in the rejected directory
    try:
        files = os.listdir(REJECTED_DIR)
        logger.info(f"Found {len(files)} files in {REJECTED_DIR}")
        
        if len(files) == 0:
            logger.info("No files found in rejected directory")
            return results
        
        for fname in sorted(files):
            if not fname.lower().endswith((".xlsx", ".xls")):
                logger.debug(f"Skipping non-Excel file: {fname}")
                continue
                
            fpath = os.path.join(REJECTED_DIR, fname)
            file_info = {"filename": fname, "sheets": {}}
            logger.info(f"Processing file: {fname}")
            
            # Try to get file size for debugging
            try:
                file_size = os.path.getsize(fpath)
                logger.info(f"File size: {file_size} bytes")
            except Exception as e:
                logger.warning(f"Could not get file size: {e}")
            
            # Strategy 1: openpyxl (standard)
            try:
                file_info["sheets"] = _read_with_openpyxl(fpath)
                results.append(file_info)
                logger.info(f"Successfully processed {fname} with openpyxl")
                continue
            except Exception as e:
                logger.warning(f"openpyxl failed for {fname}: {str(e)}")
            
            # Strategy 2: pandas (handles some XML quirks openpyxl rejects)
            try:
                file_info["sheets"] = _read_with_pandas(fpath)
                results.append(file_info)
                logger.info(f"Successfully processed {fname} with pandas")
                continue
            except Exception as e:
                logger.warning(f"pandas failed for {fname}: {str(e)}")
            
            # Strategy 3: raw zip (sheet names only, no column info)
            try:
                file_info["sheets"] = _read_as_zip(fpath)
                file_info["partial_read"] = True
                results.append(file_info)
                logger.warning(f"Partially processed {fname} as zip archive")
                continue
            except Exception as e:
                file_info["read_error"] = (
                    "File is unreadable by all methods — it may be corrupted, "
                    "password-protected, or saved in an unsupported format. "
                    f"Detail: {e}"
                )
                results.append(file_info)
                logger.error(f"Failed to process {fname}: {str(e)}")
                
    except Exception as e:
        logger.error(f"Error reading rejected directory: {str(e)}")
    
    logger.info(f"Total processed files: {len(results)}")
    return results


# Endpoint: analyze rejected files using the LLM and return guidance
@router.get("/ai/analyze-rejected")
async def analyze_rejected():
    """Analyze rejected Excel files and provide AI-powered feedback"""
    try:
        rejected_files = read_rejected_files()
        
        if not rejected_files:
            return {
                "total_rejected": 0,
                "analysis": "No rejected files found in the rejected folder. Great job! All files passed validation.",
                "rejected_directory": REJECTED_DIR
            }
        
        file_summaries = []
        for f in rejected_files:
            if "read_error" in f:
                file_summaries.append(
                    f"- {f['filename']}: UNREADABLE — {f['read_error']}"
                )
            elif f.get("partial_read"):
                sheet_names = list(f["sheets"].keys())
                file_summaries.append(
                    f"- {f['filename']}: Partially readable. "
                    f"Sheets detected: {sheet_names}. Column headers could not be extracted."
                )
            else:
                sheet_lines = []
                for sheet, cols in f["sheets"].items():
                    if cols:
                        col_str = ", ".join(cols[:20])  # Limit to first 20 columns
                        if len(cols) > 20:
                            col_str += f" ... and {len(cols) - 20} more"
                    else:
                        col_str = "(empty — no headers found)"
                    sheet_lines.append(f"    Sheet '{sheet}': {col_str}")
                body = "\n".join(sheet_lines) if sheet_lines else "    (no sheets found)"
                file_summaries.append(f"- {f['filename']}:\n{body}")
        
        files_block = "\n".join(file_summaries)
        
        prompt = f"""You are a data quality assistant for a CRM/sales ETL pipeline.

The following Excel files were REJECTED during the ETL validation process.
Analyze each file and explain exactly why it was rejected based on the rules below.

=== BUSINESS RULES ===
{BUSINESS_RULES}

=== REJECTED FILES (sheet names and column headers actually found in each file) ===
{files_block}

=== YOUR TASK ===
For each rejected file:
1. State the exact reason(s) it failed (missing sheet, missing required columns, empty sheet, or unreadable file)
2. List specifically which sheets or columns are missing by name
3. Give a clear, practical fix instruction for the person who uploaded the file

If multiple files share the same problem, group them together.
For UNREADABLE files, explain that the file itself is corrupted or in the wrong format,
and tell the user to re-export it from the source CRM system as a proper .xlsx file.
Be concise and direct. The audience is a business user, not a developer.
Response language: English
"""
        
        logger.info(f"Sending request to Ollama at {OLLAMA_URL} with model {OLLAMA_MODEL}")
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.post(
                OLLAMA_URL, 
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "num_predict": 1000,
                        "temperature": 0.7
                    }
                }
            )
            response.raise_for_status()
            result = response.json()
        
        logger.info("Successfully received response from Ollama")
        
        return {
            "total_rejected": len(rejected_files),
            "analysis": result.get("response", "No response from model."),
            "files_analyzed": [f["filename"] for f in rejected_files],
            "rejected_directory": REJECTED_DIR
        }
        
    except httpx.TimeoutException:
        logger.error("Ollama request timed out")
        raise HTTPException(
            status_code=504,
            detail="AI analysis timed out. The Ollama service might be slow or unavailable."
        )
    except httpx.HTTPStatusError as e:
        logger.error(f"Ollama HTTP error: {str(e)}")
        raise HTTPException(
            status_code=502,
            detail=f"AI service error: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Unexpected error in analyze_rejected: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Internal server error: {str(e)}"
        )


# AI service health-check endpoint to verify dependencies and connectivity
@router.get("/ai/health")
async def ai_health_check():
    """Health check endpoint for AI analysis service"""
    try:
        # Check if rejected directory exists
        dir_exists = os.path.exists(REJECTED_DIR)
        dir_writable = os.access(REJECTED_DIR, os.W_OK) if dir_exists else False
        
        # Get directory contents for debugging
        dir_contents = []
        if dir_exists:
            try:
                dir_contents = os.listdir(REJECTED_DIR)[:10]  # First 10 files
            except Exception as e:
                dir_contents = [f"Error listing directory: {str(e)}"]
        
        # Check Ollama connectivity
        ollama_healthy = False
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                response = await client.get(f"{_ollama_base.rstrip('/')}/api/tags")
                ollama_healthy = response.status_code == 200
        except Exception as e:
            logger.warning(f"Ollama health check failed: {str(e)}")
        
        return {
            "status": "healthy" if ollama_healthy and dir_exists else "degraded",
            "rejected_directory_exists": dir_exists,
            "rejected_directory_writable": dir_writable,
            "rejected_directory_path": REJECTED_DIR,
            "rejected_directory_contents": dir_contents,
            "ollama_healthy": ollama_healthy,
            "ollama_url": _ollama_base,
            "ollama_model": OLLAMA_MODEL,
            "working_directory": os.getcwd(),
            "python_version": sys.version
        }
    except Exception as e:
        logger.error(f"Health check error: {str(e)}")
        return {
            "status": "unhealthy",
            "error": str(e)
        }