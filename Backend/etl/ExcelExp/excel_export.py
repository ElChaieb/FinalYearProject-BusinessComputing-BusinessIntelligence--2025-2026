"""
excel_export.py
---------------
Standalone script that reads one or more CRM Excel files (same format as the
ETL pipeline expects) and produces a single output Excel file containing:

  Sheet         Source / Logic
  ──────────── ──────────────────────────────────────────────────────────────
  Users         Unique users from OP (Oppo_AssignedUserId) + DEVIS (User_UserId)
  Clients       Unique clients from OP (Comp_Name / email / city), auto-id'd
  Opportunities OP sheet rows, linked to client_id
  Vehicles      Unique vehicles from DEVIS, enriched with category + base_price
  Quotes        DEVIS rows filtered to opportunities that exist, with price
  Sales         Fake sales for 10–23 % of quotes (same rule as the ETL pipeline)

Usage
-----
  # Single file
  python excel_export.py path/to/CRM-AGENCY.xlsx

  # All .xlsx files in a directory
  python excel_export.py path/to/raw/

  # Explicit output path
  python excel_export.py path/to/raw/ --output my_export.xlsx

Dependencies: pandas, openpyxl  (pip install pandas openpyxl)
"""

import argparse
import os
import random
import re
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd

# ── Vehicle catalogue (copied verbatim from parse_vehicles.py) ───────────────
_CATALOGUE: dict[str, tuple[str, float]] = {
    ar_ref: (category, float(price))
    for ar_ref, category, price in [
        ("SX11-GF+-080",    "SUV Compact", 105000),
        ("SX11-GS-080",     "SUV Compact",  98000),
        ("NL-3B-A10",       "SUV",         110000),
        ("NL-3B-B07",       "SUV",         110000),
        ("SX11-GF+-F36",    "SUV Compact", 105000),
        ("NL-3B-C19",       "SUV",         110000),
        ("NL-3B-E29",       "SUV",         110000),
        ("SX11-GF+-E32",    "SUV Compact", 105000),
        ("SX11-GS-D14",     "SUV Compact",  98000),
        ("NL-3B-D06",       "SUV",         110000),
        ("GX3-CVT-OR",      "Citadine",     70000),
        ("GX3-CVT-BAS",     "Citadine",     70000),
        ("GX3-CVT-WAA",     "Citadine",     70000),
        ("SX11-GS-E32",     "SUV Compact",  98000),
        ("LF7154B-208",     "Citadine",     75000),
        ("LF7154B-C23",     "Citadine",     75000),
        ("LF7154B-K22",     "Citadine",     75000),
        ("LF7154B-E43",     "Citadine",     75000),
        ("LF7154B-G65",     "Citadine",     75000),
        ("E22H-ADA",        "Électrique",  130000),
        ("SX11-GF+-D14",    "SUV Compact", 105000),
        ("SX11-GF+-G66",    "SUV Compact", 105000),
        ("SX11-GS-G66",     "SUV Compact",  98000),
        ("SS11-BVA",        "Berline",      85000),
        ("FX11",            "SUV",         140000),
        ("FY11",            "SUV",         135000),
        ("LP5SEF",          "Électrique",  125000),
        ("KX11",            "SUV",         150000),
        ("FY11-BA-ADE",     "SUV",         138000),
        ("GX3P-MT-WAA",     "Citadine",     72000),
        ("KX11-BA-LAK",     "SUV",         150000),
        ("SX11-MT-WAA",     "SUV Compact",  95000),
        ("GX3-CVT-BAS-CRM", "Citadine",    70000),
        ("GX3P-CVT-WAA",    "Citadine",     75000),
        ("KX11-BA-ACT",     "SUV",         150000),
        ("GE13-LAK",        "Électrique",  128000),
        ("FY11-BA-LAK",     "SUV",         138000),
        ("LF7154B-MT-208",  "Citadine",     72000),
        ("LF7154B-MT-C23",  "Citadine",     72000),
        ("LF7154B-MT-G65",  "Citadine",     72000),
        ("LF7154B-MT-E43",  "Citadine",     72000),
        ("GE13-BAS",        "Électrique",  128000),
        ("SS11-BVM",        "Berline",      80000),
        ("GE13-WAA",        "Électrique",  128000),
        ("SS11-MT-WAA",     "Berline",      80000),
        ("GX3-CVT-RAM",     "Citadine",     70000),
        ("LF7154B-MT-K22",  "Citadine",     72000),
        ("SX11-GS-F36",     "SUV Compact",  98000),
        ("SS11-MT-SAI",     "Berline",      80000),
        ("FY11-BA-WAA",     "SUV",         138000),
        ("EX5-EL-ADA",      "Électrique",  132000),
        ("EX5-EL-GRE",      "Électrique",  132000),
        ("EX5-EL-LAK",      "Électrique",  132000),
        ("EX5-EL-SAI",      "Électrique",  132000),
        ("EX5-EL-WAA",      "Électrique",  132000),
        ("SX11-GF-ADA",     "SUV Compact", 100000),
        ("SX11-GF-WAA",     "SUV Compact", 100000),
        ("SS11T-MT-ADA",    "Berline",      78000),
        ("SS11T-MT-BAS",    "Berline",      78000),
        ("SS11T-MT-SAI",    "Berline",      78000),
        ("SS11T-MT-WAA",    "Berline",      78000),
        ("SS11-AT-OR",      "Berline",      85000),
        ("SX11-GS-WAA",     "SUV Compact",  98000),
        ("SX11-GF-CS",      "SUV Compact", 102000),
        ("GX3-CVT-WAA-CRM", "Citadine",    70000),
        ("GX3-MT-BAS",      "Citadine",     68000),
        ("GX3-MT-WAA",      "Citadine",     68000),
        ("KX11-BA-WAA",     "SUV",         150000),
        ("SX11-CVT-WAA",    "SUV Compact",  98000),
        ("SS11-MT-ADA",     "Berline",      80000),
        ("SS11-MT-BAS",     "Berline",      80000),
        ("EX5",             "Électrique",  132000),
        ("SS11-AT-WAA",     "Berline",      85000),
        ("GC6M72M-WAA",     "Berline",      75000),
        ("GC6M72M-LAK",     "Berline",      75000),
        ("GC6M72M-RAM",     "Berline",      75000),
        ("GC6M72M-BAS",     "Berline",      75000),
        ("GC6M72M-SAI",     "Berline",      75000),
        ("GC6M72M-OR",      "Berline",      75000),
        ("FE-3JCM72-WAA",   "Berline",      68000),
        ("GX3-MT-RAM",      "Citadine",     68000),
        ("GX3-MT-OR",       "Citadine",     68000),
        ("GX3-MT-OR-CRM",   "Citadine",     68000),
        ("GX3-MT-WAA-CRM",  "Citadine",     68000),
        ("GX3-MT-RAM-CRM",  "Citadine",     68000),
        ("GX3-CVT-OR-CRM",  "Citadine",     70000),
        ("GX3-CVT-RAM-CRM", "Citadine",     70000),
        ("GX3-MT-BAS-CRM",  "Citadine",     68000),
        ("SX11",            "SUV Compact",  95000),
        ("SS11-AT-ADA",     "Berline",      85000),
        ("SS11-AT-SAI",     "Berline",      85000),
        ("SS11-AT-BAS",     "Berline",      85000),
        ("KX11-BA-SAF",     "SUV",         150000),
        ("GE13-SAF",        "Électrique",  128000),
        ("SX11-GS-SAF",     "SUV Compact",  98000),
        ("SX11-GF-RAM",     "SUV Compact", 100000),
        ("EX5EM-I-ADA",     "Électrique",  136000),
        ("NL-3B",           "SUV",         108000),
        ("SX11-DCT-WAA",    "SUV Compact",  98000),
        ("SX11-GS-ADA",     "SUV Compact",  98000),
        ("FE-7A74-BAS",     "Berline",      72000),
        ("SX11-GS-BAS",     "SUV Compact",  98000),
    ]
}

FAKE_SALE_RATE_MIN = 0.10
FAKE_SALE_RATE_MAX = 0.23


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_int(val):
    try:
        return int(val) if val is not None and not (
            isinstance(val, float) and pd.isna(val)
        ) else None
    except (ValueError, TypeError):
        return None


def _clean(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip().title() or None


def _clean_lower(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip().lower() or None


def _clean_upper(val) -> str | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip().upper() or None


def _lookup_vehicle(ar_ref: str) -> tuple[str, float | None]:
    entry = _CATALOGUE.get(ar_ref)
    if entry:
        return entry
    for key, val in _CATALOGUE.items():
        if ar_ref.startswith(key + "-") or key.startswith(ar_ref + "-"):
            return val
    return "Inconnu", None


def _random_date_after(start: date | None, end: date) -> date:
    if start is None:
        return end
    offset = random.randint(1, 6)
    return min(start + timedelta(days=offset), end)


def _extract_agency_name(filepath: str) -> str:
    basename = Path(filepath).stem
    match = re.search(r"-(.+)$", basename)
    return match.group(1).strip().title() if match else "Unknown"


def _load_sheets(filepath: str) -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
    """Load OP and DEVIS sheets, trying calamine then openpyxl."""
    result = {}
    for sheet in ("OP", "DEVIS"):
        try:
            xl = pd.ExcelFile(filepath, engine="calamine")
            if sheet in xl.sheet_names:
                df = xl.parse(sheet)
                result[sheet] = df if not df.empty else None
            else:
                result[sheet] = None
        except Exception:
            try:
                import openpyxl.styles.fills as _fills
                _orig = _fills._assign_position

                def _patched(stops):
                    seen, out = {}, []
                    for s in (stops or []):
                        if s.position not in seen:
                            seen[s.position] = True
                            out.append(s)
                    return out

                _fills._assign_position = _patched
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filepath, read_only=True, data_only=True)
                    if sheet not in wb.sheetnames:
                        result[sheet] = None
                        continue
                    ws = wb[sheet]
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                finally:
                    _fills._assign_position = _orig

                if not rows:
                    result[sheet] = None
                    continue
                headers = [str(c) if c is not None else f"col_{i}"
                           for i, c in enumerate(rows[0])]
                df = pd.DataFrame(rows[1:], columns=headers)
                result[sheet] = df if not df.empty else None
            except Exception as e:
                print(f"  [WARN] Could not load sheet '{sheet}': {e}", file=sys.stderr)
                result[sheet] = None

    return result.get("OP"), result.get("DEVIS")


# ── Per-sheet extractors ──────────────────────────────────────────────────────

def extract_users(df_op: pd.DataFrame, df_devis: pd.DataFrame | None) -> dict[int, dict]:
    """Returns {user_id: user_dict}"""
    users: dict[int, dict] = {}

    if df_op is not None:
        for _, row in df_op.iterrows():
            uid = _safe_int(row.get("Oppo_AssignedUserId"))
            if uid is None or uid in users:
                continue
            users[uid] = {
                "user_id":    uid,
                "last_name":  _clean(row.get("User_LastName")),
                "first_name": _clean(row.get("User_FirstName")),
                "email":      _clean_lower(row.get("User_EmailAddress")),
                "role":       "Commercial",
                "agency_name": _clean(row.get("Chan_Description")),
            }

    if df_devis is not None and "User_UserId" in df_devis.columns:
        for _, row in df_devis.iterrows():
            uid = _safe_int(row.get("User_UserId"))
            if uid is None or uid in users:
                continue
            users[uid] = {
                "user_id":    uid,
                "last_name":  _clean(row.get("User_LastName")),
                "first_name": _clean(row.get("User_FirstName")),
                "email":      None,
                "role":       "Commercial",
                "agency_name": None,
            }

    return users


def extract_clients(df_op: pd.DataFrame) -> tuple[list[dict], dict]:
    """
    Returns (clients_list, name_email_to_id mapping).
    client_id is assigned as a local auto-increment (no DB).
    """
    clients: list[dict] = []
    lookup: dict[tuple, int] = {}   # (upper_name, lower_email) → client_id
    next_id = 1

    if df_op is None:
        return clients, {}

    for _, row in df_op.iterrows():
        name_raw  = row.get("Comp_Name")
        email_raw = row.get("Emai_EmailAddress")
        city_raw  = row.get("Addr_City")

        name_norm  = _clean_upper(name_raw)
        email_norm = _clean_lower(email_raw)

        if not name_norm:
            continue

        key = (name_norm, email_norm)
        if key in lookup:
            continue

        client_id = next_id
        next_id += 1
        lookup[key] = client_id
        clients.append({
            "client_id": client_id,
            "full_name": name_norm.title(),
            "email":     email_norm,
            "city":      _clean(city_raw),
        })

    return clients, lookup


def extract_opportunities(df_op: pd.DataFrame, client_lookup: dict) -> list[dict]:
    if df_op is None:
        return []

    df = (
        df_op
        .dropna(subset=["Oppo_OpportunityId"])
        .drop_duplicates(subset=["Oppo_OpportunityId"])
        .copy()
    )
    df["Oppo_CreatedDate"] = pd.to_datetime(
        df["Oppo_CreatedDate"], dayfirst=True, errors="coerce"
    )

    records = []
    for _, row in df.iterrows():
        oppo_id = _safe_int(row["Oppo_OpportunityId"])
        if oppo_id is None:
            continue

        name_norm  = _clean_upper(row.get("Comp_Name"))
        email_norm = _clean_lower(row.get("Emai_EmailAddress"))
        client_id  = client_lookup.get((name_norm, email_norm))

        raw_date     = row.get("Oppo_CreatedDate")
        created_date = (
            raw_date.date()
            if pd.notna(raw_date) and hasattr(raw_date, "date")
            else None
        )

        records.append({
            "oppo_id":          oppo_id,
            "user_id":          _safe_int(row.get("Oppo_AssignedUserId")),
            "client_id":        client_id,
            "agency_name":      _clean(row.get("Chan_Description")),
            "created_date":     created_date,
            "client_reference": _clean(row.get("comp_reference")),
            "deleted":          True,   # will be flipped after quotes processed
        })
    return records


def extract_vehicles(df_devis: pd.DataFrame) -> tuple[list[dict], dict[str, float | None]]:
    """Returns (vehicles_list, ar_ref_to_price mapping)"""
    if df_devis is None or "AR_Ref" not in df_devis.columns:
        return [], {}

    keep = [c for c in ("AR_Ref", "AR_Design", "Marque", "Modèle") if c in df_devis.columns]
    vehicles_df = (
        df_devis[keep]
        .dropna(subset=["AR_Ref"])
        .drop_duplicates(subset=["AR_Ref"])
    )

    vehicles = []
    price_map: dict[str, float | None] = {}
    for _, row in vehicles_df.iterrows():
        ar_ref = str(row["AR_Ref"]).strip()
        category, base_price = _lookup_vehicle(ar_ref)
        price_map[ar_ref] = base_price
        vehicles.append({
            "ar_ref":      ar_ref,
            "ar_design":   _clean(row.get("AR_Design")),
            "brand":       _clean(row.get("Marque")),
            "model":       _clean(row.get("Modèle")),
            "category":    category,
            "base_price":  base_price,
        })
    return vehicles, price_map


def extract_quotes_and_sales(
    df_devis: pd.DataFrame,
    oppo_set: set[int],           # oppo_ids that exist in opportunities
    oppo_agency: dict[int, str],  # oppo_id → agency_name
    oppo_client: dict[int, int],  # oppo_id → client_id
    price_map: dict[str, float | None],
    users: dict[int, dict],
) -> tuple[list[dict], list[dict]]:
    """
    Returns (quotes_list, sales_list).
    Mirrors the ETL logic exactly:
      - Only quotes with a valid oppo_id in oppo_set are kept.
      - One quote per opportunity (first wins).
      - 10–23% of inserted quotes become fake sales.
      - Won quotes have deleted=False; losers have deleted=True.
    """
    if df_devis is None or "Quot_OrderQuoteID" not in df_devis.columns:
        return [], []

    df = (
        df_devis
        .dropna(subset=["Quot_OrderQuoteID"])
        .drop_duplicates(subset=["Quot_OrderQuoteID"], keep="first")
        .copy()
    )
    if "Quot_CreatedDate" in df.columns:
        df["Quot_CreatedDate"] = pd.to_datetime(
            df["Quot_CreatedDate"], dayfirst=True, errors="coerce"
        )
    else:
        df["Quot_CreatedDate"] = pd.NaT

    has_user_id = "User_UserId" in df.columns

    # name→id fallback cache
    name_to_uid: dict[tuple, int | None] = {}

    def resolve_user(row) -> int | None:
        if has_user_id:
            uid = _safe_int(row.get("User_UserId"))
            return uid if uid in users else None
        ln = _clean_upper(row.get("User_LastName"))
        fn = _clean_upper(row.get("User_FirstName"))
        key = (ln, fn)
        if key not in name_to_uid:
            found = None
            for u in users.values():
                if (
                    u.get("last_name", "").upper() == (ln or "")
                    and u.get("first_name", "").upper() == (fn or "")
                ):
                    found = u["user_id"]
                    break
            name_to_uid[key] = found
        return name_to_uid[key]

    oppo_claimed: set[int] = set()
    newly_inserted: list[dict] = []

    for _, row in df.iterrows():
        quote_id = _safe_int(row["Quot_OrderQuoteID"])
        if quote_id is None:
            continue

        oppo_id = _safe_int(row.get("Quot_opportunityid"))
        if oppo_id is None or oppo_id not in oppo_set:
            continue
        if oppo_id in oppo_claimed:
            continue

        ar_ref = _clean(row.get("AR_Ref"))
        if ar_ref:
            ar_ref = str(row.get("AR_Ref")).strip()  # preserve case
        price = price_map.get(ar_ref) if ar_ref else None

        raw_date     = row.get("Quot_CreatedDate")
        created_date = (
            raw_date.date()
            if pd.notna(raw_date) and hasattr(raw_date, "date")
            else None
        )

        user_id     = resolve_user(row)
        agency_name = oppo_agency.get(oppo_id)
        client_id   = oppo_client.get(oppo_id)

        newly_inserted.append({
            "quote_id":     quote_id,
            "oppo_id":      oppo_id,
            "ar_ref":       ar_ref,
            "user_id":      user_id,
            "client_id":    client_id,
            "agency_name":  agency_name,
            "price":        price,
            "created_date": created_date,
            "deleted":      True,    # default; flipped for winners below
        })
        oppo_claimed.add(oppo_id)

    # ── Fake sales ────────────────────────────────────────────────────────────
    sales: list[dict] = []
    won_ids: set[int] = set()

    if newly_inserted:
        rate = random.uniform(FAKE_SALE_RATE_MIN, FAKE_SALE_RATE_MAX)
        k    = round(len(newly_inserted) * rate)
        if k > 0:
            today  = date.today()
            chosen = random.sample(newly_inserted, min(k, len(newly_inserted)))
            for q in chosen:
                sale_date = _random_date_after(q["created_date"], today)
                sales.append({
                    "quote_id":    q["quote_id"],
                    "oppo_id":     q["oppo_id"],
                    "user_id":     q["user_id"],
                    "client_id":   q["client_id"],
                    "ar_ref":      q["ar_ref"],
                    "agency_name": q["agency_name"],
                    "sale_date":   sale_date,
                    "quantity":    1,
                    "final_price": q["price"],
                })
                won_ids.add(q["quote_id"])

    # flip won quotes to deleted=False
    quotes = []
    for q in newly_inserted:
        q["deleted"] = q["quote_id"] not in won_ids
        quotes.append(q)

    return quotes, sales


# ── Core: process one file ────────────────────────────────────────────────────

def process_file(filepath: str) -> dict:
    """Extract all data from one xlsx file and return as dict of lists."""
    print(f"Processing: {Path(filepath).name}")

    df_op, df_devis = _load_sheets(filepath)
    agency_name = _extract_agency_name(filepath)

    print(f"  OP rows:    {len(df_op) if df_op is not None else 'MISSING'}")
    print(f"  DEVIS rows: {len(df_devis) if df_devis is not None else 'MISSING'}")

    # Users
    users = extract_users(df_op, df_devis)
    print(f"  Users found: {len(users)}")

    # Clients
    clients, client_lookup = extract_clients(df_op)
    print(f"  Clients found: {len(clients)}")

    # Opportunities
    opportunities = extract_opportunities(df_op, client_lookup)
    print(f"  Opportunities found: {len(opportunities)}")

    # Vehicles
    vehicles, price_map = extract_vehicles(df_devis)
    print(f"  Vehicles found: {len(vehicles)}")

    # Build lookup maps for quotes
    oppo_set    = {o["oppo_id"] for o in opportunities}
    oppo_agency = {o["oppo_id"]: o["agency_name"] for o in opportunities}
    oppo_client = {o["oppo_id"]: o["client_id"]   for o in opportunities}

    # Quotes + fake sales
    quotes, sales = extract_quotes_and_sales(
        df_devis, oppo_set, oppo_agency, oppo_client, price_map, users
    )
    print(f"  Quotes inserted: {len(quotes)}, Sales generated: {len(sales)}")

    # Mark won opportunities (those that have at least one quote)
    quoted_oppos = {q["oppo_id"] for q in quotes}
    for o in opportunities:
        if o["oppo_id"] in quoted_oppos:
            o["deleted"] = False

    return {
        "users":         list(users.values()),
        "clients":       clients,
        "opportunities": opportunities,
        "vehicles":      vehicles,
        "quotes":        quotes,
        "sales":         sales,
    }


# ── Merge results from multiple files ────────────────────────────────────────

def merge_results(all_results: list[dict]) -> dict:
    """
    Merge data across multiple files.
    Deduplication keys:
      users         → user_id
      clients       → (upper full_name, lower email)    (re-assign sequential IDs)
      opportunities → oppo_id
      vehicles      → ar_ref
      quotes        → quote_id
      sales         → quote_id
    """
    # Users
    seen_users: dict[int, dict] = {}
    for r in all_results:
        for u in r["users"]:
            seen_users.setdefault(u["user_id"], u)

    # Clients — re-assign IDs across files
    seen_clients: dict[tuple, dict] = {}
    next_cid = 1
    old_to_new: list[dict[int, int]] = []   # per-file old→new client_id mapping

    for r in all_results:
        mapping: dict[int, int] = {}
        for c in r["clients"]:
            key = (_clean_upper(c["full_name"]), _clean_lower(c["email"]))
            if key not in seen_clients:
                new_c = dict(c)
                new_c["client_id"] = next_cid
                seen_clients[key] = new_c
                next_cid += 1
            mapping[c["client_id"]] = seen_clients[key]["client_id"]
        old_to_new.append(mapping)

    # Remap client_ids in opportunities / quotes / sales
    def remap(record: dict, mapping: dict[int, int]) -> dict:
        r2 = dict(record)
        if "client_id" in r2 and r2["client_id"] is not None:
            r2["client_id"] = mapping.get(r2["client_id"], r2["client_id"])
        return r2

    seen_oppos:    dict[int, dict] = {}
    seen_vehicles: dict[str, dict] = {}
    seen_quotes:   dict[int, dict] = {}
    seen_sales:    dict[int, dict] = {}

    for i, r in enumerate(all_results):
        mapping = old_to_new[i]
        for o in r["opportunities"]:
            seen_oppos.setdefault(o["oppo_id"], remap(o, mapping))
        for v in r["vehicles"]:
            seen_vehicles.setdefault(v["ar_ref"], v)
        for q in r["quotes"]:
            seen_quotes.setdefault(q["quote_id"], remap(q, mapping))
        for s in r["sales"]:
            seen_sales.setdefault(s["quote_id"], remap(s, mapping))

    return {
        "users":         list(seen_users.values()),
        "clients":       list(seen_clients.values()),
        "opportunities": list(seen_oppos.values()),
        "vehicles":      list(seen_vehicles.values()),
        "quotes":        list(seen_quotes.values()),
        "sales":         list(seen_sales.values()),
    }


# ── Write output Excel ────────────────────────────────────────────────────────

def write_excel(data: dict, output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sheet_order = [
            ("Users",         data["users"]),
            ("Clients",       data["clients"]),
            ("Opportunities", data["opportunities"]),
            ("Vehicles",      data["vehicles"]),
            ("Quotes",        data["quotes"]),
            ("Sales",         data["sales"]),
        ]
        for sheet_name, records in sheet_order:
            df = pd.DataFrame(records) if records else pd.DataFrame()
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  {sheet_name}: {len(records)} rows")

    print(f"\nOutput written to: {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def collect_input_files(path: str) -> list[str]:
    p = Path(path)
    if p.is_dir():
        files = sorted(
            str(f) for f in p.glob("*.xlsx") if not f.name.startswith("~$")
        )
        if not files:
            print(f"No .xlsx files found in {path}", file=sys.stderr)
            sys.exit(1)
        return files
    if p.is_file() and p.suffix == ".xlsx":
        return [str(p)]
    print(f"Invalid input: {path}", file=sys.stderr)
    sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export CRM Excel files → structured Excel workbook"
    )
    parser.add_argument(
        "input",
        help="Path to a single .xlsx file or a directory of .xlsx files",
    )
    parser.add_argument(
        "--output", "-o",
        default="crm_export.xlsx",
        help="Output file path (default: crm_export.xlsx)",
    )
    args = parser.parse_args()

    files = collect_input_files(args.input)
    print(f"Found {len(files)} file(s) to process\n")

    all_results = [process_file(f) for f in files]
    merged      = merge_results(all_results) if len(all_results) > 1 else all_results[0]

    print(f"\nWriting output Excel ({args.output})...")
    write_excel(merged, args.output)


if __name__ == "__main__":
    main()
