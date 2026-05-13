"""
generate_fake_sales.py
─────────────────────
Standalone script that reads one or more CRM Excel files (each with OP + DEVIS
sheets), generates fake sales for 75–85 % of valid quotes, and writes a new
Excel file per input into an `output/` folder next to the source files.

The generated file is a clean copy of the original with an added SALES sheet:
  quote_id | oppo_id | ar_ref | user_id | sale_date | quantity | final_price

The ETL pipeline then processes the generated file normally. No fake-data
logic is needed inside the pipeline itself.

Usage
─────
  # Process all xlsx files in a folder
  python generate_fake_sales.py path/to/raw/

  # Process specific files
  python generate_fake_sales.py file1.xlsx file2.xlsx

  # Explicit output folder
  python generate_fake_sales.py path/to/raw/ --output path/to/out/
"""

import argparse
import os
import random
import shutil
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Vehicle catalogue (ar_ref → base_price) ──────────────────────────────────
_CATALOGUE: dict[str, float] = {
    ar_ref: float(price)
    for ar_ref, price in [
        ("SX11-GF+-080",    105000), ("SX11-GS-080",      98000),
        ("NL-3B-A10",       110000), ("NL-3B-B07",        110000),
        ("SX11-GF+-F36",    105000), ("NL-3B-C19",        110000),
        ("NL-3B-E29",       110000), ("SX11-GF+-E32",     105000),
        ("SX11-GS-D14",      98000), ("NL-3B-D06",        110000),
        ("GX3-CVT-OR",       70000), ("GX3-CVT-BAS",       70000),
        ("GX3-CVT-WAA",      70000), ("SX11-GS-E32",       98000),
        ("LF7154B-208",      75000), ("LF7154B-C23",        75000),
        ("LF7154B-K22",      75000), ("LF7154B-E43",        75000),
        ("LF7154B-G65",      75000), ("E22H-ADA",          130000),
        ("SX11-GF+-D14",    105000), ("SX11-GF+-G66",     105000),
        ("SX11-GS-G66",      98000), ("SS11-BVA",           85000),
        ("FX11",            140000), ("FY11",              135000),
        ("LP5SEF",          125000), ("KX11",              150000),
        ("FY11-BA-ADE",     138000), ("GX3P-MT-WAA",        72000),
        ("KX11-BA-LAK",     150000), ("SX11-MT-WAA",        95000),
        ("GX3-CVT-BAS-CRM",  70000), ("GX3P-CVT-WAA",       75000),
        ("KX11-BA-ACT",     150000), ("GE13-LAK",          128000),
        ("FY11-BA-LAK",     138000), ("LF7154B-MT-208",     72000),
        ("LF7154B-MT-C23",   72000), ("LF7154B-MT-G65",     72000),
        ("LF7154B-MT-E43",   72000), ("GE13-BAS",          128000),
        ("SS11-BVM",         80000), ("GE13-WAA",          128000),
        ("SS11-MT-WAA",      80000), ("GX3-CVT-RAM",        70000),
        ("LF7154B-MT-K22",   72000), ("SX11-GS-F36",        98000),
        ("SS11-MT-SAI",      80000), ("FY11-BA-WAA",       138000),
        ("EX5-EL-ADA",      132000), ("EX5-EL-GRE",        132000),
        ("EX5-EL-LAK",      132000), ("EX5-EL-SAI",        132000),
        ("EX5-EL-WAA",      132000), ("SX11-GF-ADA",       100000),
        ("SX11-GF-WAA",     100000), ("SS11T-MT-ADA",       78000),
        ("SS11T-MT-BAS",     78000), ("SS11T-MT-SAI",       78000),
        ("SS11T-MT-WAA",     78000), ("SS11-AT-OR",         85000),
        ("SX11-GS-WAA",      98000), ("SX11-GF-CS",        102000),
        ("GX3-CVT-WAA-CRM",  70000), ("GX3-MT-BAS",         68000),
        ("GX3-MT-WAA",       68000), ("KX11-BA-WAA",       150000),
        ("SX11-CVT-WAA",     98000), ("SS11-MT-ADA",        80000),
        ("SS11-MT-BAS",      80000), ("EX5",               132000),
        ("SS11-AT-WAA",      85000), ("GC6M72M-WAA",        75000),
        ("GC6M72M-LAK",      75000), ("GC6M72M-RAM",        75000),
        ("GC6M72M-BAS",      75000), ("GC6M72M-SAI",        75000),
        ("GC6M72M-OR",       75000), ("FE-3JCM72-WAA",      68000),
        ("GX3-MT-RAM",       68000), ("GX3-MT-OR",          68000),
        ("GX3-MT-OR-CRM",    68000), ("GX3-MT-WAA-CRM",     68000),
        ("GX3-MT-RAM-CRM",   68000), ("GX3-CVT-OR-CRM",     70000),
        ("GX3-CVT-RAM-CRM",  70000), ("GX3-MT-BAS-CRM",     68000),
        ("SX11",             95000), ("SS11-AT-ADA",         85000),
        ("SS11-AT-SAI",      85000), ("SS11-AT-BAS",         85000),
        ("KX11-BA-SAF",     150000), ("GE13-SAF",          128000),
        ("SX11-GS-SAF",      98000), ("SX11-GF-RAM",       100000),
        ("EX5EM-I-ADA",     136000), ("NL-3B",             108000),
        ("SX11-DCT-WAA",     98000), ("SX11-GS-ADA",        98000),
        ("FE-7A74-BAS",      72000), ("SX11-GS-BAS",        98000),
    ]
}

FAKE_SALE_RATE_MIN = 0.75
FAKE_SALE_RATE_MAX = 0.85

SALES_COLUMNS = [
    "quote_id", "oppo_id", "ar_ref", "user_id",
    "sale_date", "quantity", "final_price",
]

HEADER_FILL  = PatternFill("solid", start_color="2F5496", end_color="2F5496")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FONT     = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_int(val):
    try:
        return int(val) if val is not None and not (
            isinstance(val, float) and pd.isna(val)
        ) else None
    except (ValueError, TypeError):
        return None


def _clean(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip() or None


def _lookup_price(ar_ref: str | None) -> float | None:
    if not ar_ref:
        return None
    if ar_ref in _CATALOGUE:
        return _CATALOGUE[ar_ref]
    for key, price in _CATALOGUE.items():
        if ar_ref.startswith(key + "-") or key.startswith(ar_ref + "-"):
            return price
    return None


def _random_date_after(start: date | None, end: date) -> date:
    if start is None:
        return end
    result = start + timedelta(days=random.randint(1, 6))
    return min(result, end)


# ── Core logic ────────────────────────────────────────────────────────────────

def _build_sales(df_op: pd.DataFrame, df_devis: pd.DataFrame) -> pd.DataFrame:
    """
    Derive valid quotes from OP + DEVIS and generate fake sales for 75–85 % of them.
    Returns a DataFrame with SALES_COLUMNS ready to write as the SALES sheet.
    """
    # Valid oppo_ids from OP
    valid_oppos: set[int] = set()
    for _, row in df_op.iterrows():
        oid = _safe_int(row.get("Oppo_OpportunityId"))
        if oid is not None:
            valid_oppos.add(oid)

    if "Quot_OrderQuoteID" not in df_devis.columns:
        return pd.DataFrame(columns=SALES_COLUMNS)

    df = (
        df_devis
        .dropna(subset=["Quot_OrderQuoteID"])
        .drop_duplicates(subset=["Quot_OrderQuoteID"], keep="first")
        .copy()
    )
    if "Quot_CreatedDate" in df.columns:
        df["Quot_CreatedDate"] = pd.to_datetime(
            df["Quot_CreatedDate"], dayfirst=True, errors="coerce"
        )
    else:
        df["Quot_CreatedDate"] = pd.NaT

    # Build candidate quote records, keeping only those linked to a valid oppo
    candidates: list[dict] = []
    seen_oppos: set[int] = set()

    for _, row in df.iterrows():
        quote_id = _safe_int(row["Quot_OrderQuoteID"])
        oppo_id  = _safe_int(row.get("Quot_opportunityid"))
        if quote_id is None or oppo_id is None:
            continue
        if oppo_id not in valid_oppos:
            continue
        if oppo_id in seen_oppos:        # one quote per opportunity
            continue
        seen_oppos.add(oppo_id)

        raw_date     = row.get("Quot_CreatedDate")
        created_date = (
            raw_date.date()
            if pd.notna(raw_date) and hasattr(raw_date, "date")
            else None
        )
        ar_ref  = _clean(row.get("AR_Ref"))
        user_id = _safe_int(row.get("User_UserId"))

        candidates.append({
            "quote_id":     quote_id,
            "oppo_id":      oppo_id,
            "ar_ref":       ar_ref,
            "user_id":      user_id,
            "created_date": created_date,
            "price":        _lookup_price(ar_ref),
        })

    if not candidates:
        return pd.DataFrame(columns=SALES_COLUMNS)

    # Sample 75–85 %
    rate = random.uniform(FAKE_SALE_RATE_MIN, FAKE_SALE_RATE_MAX)
    k    = max(1, round(len(candidates) * rate))
    won  = random.sample(candidates, min(k, len(candidates)))

    today = date.today()
    rows  = []
    for q in won:
        rows.append({
            "quote_id":    q["quote_id"],
            "oppo_id":     q["oppo_id"],
            "ar_ref":      q["ar_ref"],
            "user_id":     q["user_id"],
            "sale_date":   _random_date_after(q["created_date"], today),
            "quantity":    1,
            "final_price": q["price"],
        })

    return pd.DataFrame(rows, columns=SALES_COLUMNS)


def _style_sales_sheet(ws, n_rows: int) -> None:
    """Apply header styling and alternating row fill to the SALES sheet."""
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, n_rows + 2):
        fill = ALT_FILL if r % 2 == 0 else None
        for cell in ws[r]:
            cell.font = ROW_FONT
            if fill:
                cell.fill = fill

    col_widths = {
        "A": 12, "B": 12, "C": 22, "D": 12, "E": 14, "F": 10, "G": 14,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 18


def _load_sheet_patched(src_path: Path, sheet_name: str):
    """
    Load a sheet tolerating corrupt stylesheets (same fix as the ETL pipeline).
    Returns a DataFrame or None.
    """
    # Try calamine first (fast, ignores stylesheets entirely)
    try:
        xl = pd.ExcelFile(src_path, engine="calamine")
        if sheet_name not in xl.sheet_names:
            return None
        df = xl.parse(sheet_name)
        return df if not df.empty else None
    except Exception:
        pass

    # openpyxl fallback with duplicate-gradient-stop patch
    import openpyxl.styles.fills as _fills
    _orig = _fills._assign_position

    def _patched(stops):
        seen, result = {}, []
        for s in (stops or []):
            if s.position not in seen:
                seen[s.position] = True
                result.append(s)
        return result

    _fills._assign_position = _patched
    try:
        wb = load_workbook(src_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        df = pd.DataFrame(rows[1:], columns=headers)
        return df if not df.empty else None
    finally:
        _fills._assign_position = _orig


def _process_file(src_path: Path, out_dir: Path) -> str:
    """
    Read src_path, generate fake sales, write enriched copy to out_dir.
    Returns a status string.
    """
    # ── Load sheets ──────────────────────────────────────────────────────────
    try:
        df_op    = _load_sheet_patched(src_path, "OP")
        df_devis = _load_sheet_patched(src_path, "DEVIS")
    except Exception as e:
        return f"SKIP  — could not open: {e}"

    missing = [s for s, df in (("OP", df_op), ("DEVIS", df_devis)) if df is None]
    if missing:
        return f"SKIP  — missing sheets: {missing}"

    if df_op.empty or df_devis.empty:
        return "SKIP  — OP or DEVIS sheet is empty"

    # ── Generate fake sales ──────────────────────────────────────────────────
    df_sales = _build_sales(df_op, df_devis)

    # ── Copy original file then add SALES sheet ──────────────────────────────
    out_dir.mkdir(parents=True, exist_ok=True)
    dest_path = out_dir / src_path.name
    shutil.copy2(src_path, dest_path)

    import openpyxl.styles.fills as _fills
    _orig_w = _fills._assign_position

    def _patched_w(stops):
        seen, result = {}, []
        for s in (stops or []):
            if s.position not in seen:
                seen[s.position] = True
                result.append(s)
        return result

    _fills._assign_position = _patched_w
    try:
        wb = load_workbook(dest_path)
    finally:
        _fills._assign_position = _orig_w

    # Remove existing SALES sheet if any (idempotent)
    if "SALES" in wb.sheetnames:
        del wb["SALES"]

    ws = wb.create_sheet("SALES")

    # Header row
    ws.append(SALES_COLUMNS)

    # Data rows
    for _, row in df_sales.iterrows():
        ws.append([
            row["quote_id"],
            row["oppo_id"],
            row["ar_ref"],
            row["user_id"],
            row["sale_date"].isoformat() if row["sale_date"] else None,
            row["quantity"],
            row["final_price"],
        ])

    _style_sales_sheet(ws, len(df_sales))
    wb.save(dest_path)

    rate_pct = len(df_sales) / max(1, len(df_op)) * 100
    return (
        f"OK    — {len(df_sales)} sales generated "
        f"({rate_pct:.1f}% of opportunities) → {dest_path.name}"
    )


# ── CLI ───────────────────────────────────────────────────────────────────────

def _resolve_inputs(paths: list[str]) -> list[Path]:
    files: list[Path] = []
    for p in paths:
        path = Path(p)
        if path.is_dir():
            files.extend(
                f for f in sorted(path.glob("*.xlsx"))
                if not f.name.startswith("~$")
            )
        elif path.is_file() and path.suffix.lower() == ".xlsx":
            files.append(path)
        else:
            print(f"  [WARN] not a file or folder: {p}", file=sys.stderr)
    return files


def main():
    parser = argparse.ArgumentParser(
        description="Generate fake sales Excel files ready for ETL ingestion."
    )
    parser.add_argument(
        "inputs", nargs="+",
        help="One or more .xlsx files or folders containing .xlsx files.",
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help=(
            "Output folder. Defaults to output/ inside each input file's parent "
            "directory."
        ),
    )
    args = parser.parse_args()

    files = _resolve_inputs(args.inputs)
    if not files:
        print("No .xlsx files found.", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(files)} file(s) to process.\n")

    for src in files:
        out_dir = Path(args.output) if args.output else src.parent / "output"
        status  = _process_file(src, out_dir)
        print(f"  {src.name}: {status}")

    print("\nDone.")


if __name__ == "__main__":
    main()
